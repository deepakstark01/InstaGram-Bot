import os
import random
from tkinter import *
import tkinter as tk
import sys
import time
import pyautogui
import sys
import requests
import xlsxwriter
import time
import wget
import random
import shutil
import selenium
from selenium import webdriver
from pynput.keyboard import Controller, Key
from selenium.webdriver.common.keys import Keys
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
global outWorkbook
outWorkbook = xlsxwriter.Workbook("ScrapedProfile.xlsx")
outSheet = outWorkbook.add_worksheet()
outSheet.write("A1", "UserIds")
outSheet.write("B1", "Follower")
outSheet.write("C1", "Following")
outSheet.write("D1", "Link")
global rowData
rowData=1
try:
    path = os.getcwd()
    path = os.path.join(path, "Gui\\src\\image")
    shutil.rmtree(path)
except Exception:
    pass
path = os.getcwd()
path = os.path.join(path, "Gui\\src\\image")
os.mkdir(path)

def run_chrome():
    global driver
    chrome_options = webdriver.ChromeOptions()

    chrome_options.add_argument(r'--user-data-dir=C:\Users\Deepak\AppData\Local\Google\Chrome\User Data')
    # chrome_options.add_argument(r'--user-data-dir=C:\Users\mikel\AppData\Local\Google\Chrome\User Data')
    # chrome_options.add_argument(r'--incognito')
    driver = webdriver.Chrome(executable_path="C:\chromedriver.exe", options=chrome_options)

    print("work")
    url="https://www.instagram.com/"
    driver.get(url)
try:
    run_chrome()
except Exception:
    driver.close()
    print("expt")
    run_chrome()
def convert_str_to_number(x):
    total_stars = 0
    num_map = {'K':1000, 'M':1000000, 'B':1000000000}
    if x.isdigit():
        total_stars = int(x)
    else:
        x=x.replace(",","")
        if len(x) > 1:
            total_stars = float(x[:-1]) * num_map.get(x[-1].upper(), 1)
    return int(total_stars)



def scrapePostProfile(comptitorIDs):
    tInstaIds = totalIds.get()
    privateCheck = private.get()
    maxFollower = maxFollowData.get()
    minFollwer = minFollowData.get()
    minFollowing = minFollowingData.get()
    maxFollwing = maxFollowingData.get()
    tPos = totPost.get()
    followeComment=[]


    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//body//div//div[contains(@class,'_aa')]//div//div//div[1]//div[1]//a[1]"))).click()

    nextbutton ="//div[contains(@class,'_aaqg _aaqh')]//span"
    time.sleep(2)
    for i in range(0,tPos):

        time.sleep(random.randrange(3, 6))
        html = driver.page_source
        soup1 = BeautifulSoup(html, "html.parser")
        followeDataComment = soup1.find_all('a', {"class": "oajrlxb2 g5ia77u1 qu0x051f esr5mh6w e9989ue4 r7d6kgcz rq0escxv nhd2j8a9 nc684nl6 p7hjln8o kvgmc6g5 cxmmr5t8 oygrvhab hcukyx3x jb3vyjys rz4wbd8a qt6c0cv9 a8nywdso i1ao9s8h esuyzwwr f1sip0of lzcic4wl _acan _acao _acat _acaw _a6hd"})
        followeComment.extend(list(set(followeDataComment)))
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH, nextbutton))).click()


    #     WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//body//div[contains(@role,'presentation')]//section//div//div//div//div[1]"))).click()
    #     time.sleep(3)
    #     followeDataLike = soup1.find_all('a',{"class":"notranslate"})

    for followerData in followeComment:
        global rowData
        print(followerData.text)
        if (comptitorIDs != followerData.text):

            userLink = "https://www.instagram.com/" + str(followerData.text)
            driver.get(userLink)
            time.sleep(random.randrange(4, 8))
            link = driver.current_url
            html2 = driver.page_source
            soup2 = BeautifulSoup(html2, "html.parser")
            try:
                privateOrNot = soup2.find('h2', {"class": "_aa_u"}).text
                if (str(privateOrNot) == 'This account is private'):
                    print("account is private")
                    if (privateCheck):
                        print("pravte included")
                        outSheet.write(rowData, 0, followerData.text)
                        outSheet.write(rowData, 1, numFollower)
                        outSheet.write(rowData, 2, numFollowing)
                        outSheet.write(rowData, 3, link)
                        rowData += 1
                        print("row", rowData - 1)
                    else:
                        print("Skipping account is private")

            except:
                countFollwer = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, ".//*[contains(text(), 'followers')]/span"))).text

                countFollowing = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, ".//*[contains(text(), 'following')]/span"))).text
                countFollwer = countFollwer.replace(",", "")
                countFollowing = countFollowing.replace(",", "")
                numFollowing = convert_str_to_number(countFollowing)
                numFollower = convert_str_to_number(countFollwer)
                if ((minFollowing <= numFollowing <= maxFollwing) and (minFollwer <= numFollower <= maxFollower)):
                    time.sleep(random.randrange(3, 6))
                    html3 = driver.page_source
                    soup3 = BeautifulSoup(html3, "html.parser")
                    userdataPic = soup3.find('span', {"class": "_aa8h"})
                    imglink = userdataPic.find('img', {"draggable": "false"})["src"]
                    save_as = os.path.join(path, "image"+str(rowData) + '.jpg')
                    wget.download(imglink, save_as)
                    outSheet.write(rowData, 0, followerData.text)
                    outSheet.write(rowData, 1, numFollower)
                    outSheet.write(rowData, 2, numFollowing)
                    outSheet.write(rowData, 3, link)
                    rowData += 1
                    print("row", rowData - 1,followerData.text,": ",numFollower,"follower :",numFollowing,"Following")
            if (rowData == tInstaIds):
                print("done Scrapping!!!!")
                outWorkbook.close()
                quit()

        time.sleep(3)

def excelFollow():
    # get all data in sheet

    global rowData
    file = "comptitor.xlsx"
    wb_obj = openpyxl.load_workbook(file)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    timeValueSleep = timeDelayFollow.get()
    for i in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)
        Insta_Id = cell_obj.value
        Insta_Id_Urls = "https://www.instagram.com/" + Insta_Id
        print(str(i) + ":" + Insta_Id_Urls)
        driver.get(Insta_Id_Urls)
        profiles = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "*//h2[@class=class=_7UhW9]"))).text
        print(profiles)
        scrapePostProfile(Insta_Id)
        print(rowData)
        time.sleep(timeValueSleep)

def run_All():
    # run_chrome()
    excelFollow()
    try:
        outWorkbook.close()
    except Exception:
        pass
    driver.close()


root = Tk()
root.geometry("900x400")
root.title("Author Darakhsha_Deepak")
root['background'] = '#008000'
off_color = "red"
on_color = "green"
# credential
private =IntVar()
Label(root, text="min follower", font="Roboto,14,bold", bg="#008000", fg="black").grid(row=0,column=0)
Label(root, text="max follower", font="Roboto,14,bold", bg="#008000", fg="black").grid(row=0,column=2)
Label(root, text="min Following ", font="Roboto,14,bold", bg="#008000", fg="black").grid(row=1,column=0)
Label(root, text="max Following ", font="Roboto,14,bold", bg="#008000", fg="black").grid(row=1,column=2)
Label(root, text="Time delay", font="Roboto,14,bold", bg="#008000", fg="black").grid(row=2,column=0)
Label(root, text="total id's want scrape", font="Roboto,14,bold", bg="#008000", fg="black").grid(row=2,column=2)
Label(root, text="Number of post", font="Roboto,14,bold", bg="#008000", fg="black").grid(row=3,column=0)
Label(root, text="Private Or Not", font="Roboto,14,bold", bg="#008000", fg="black").grid(row=4, column=0)
Checkbutton(root, text="Private", variable=private).grid(row=5 ,column=0)
# checkbox for private , photo
# daily follow, unfollow
#like follow


# password.grid(row=1)

# entry
maxFollowData= IntVar()
minFollowData= IntVar()
maxFollowingData= IntVar()
minFollowingData= IntVar()
totalIds =IntVar()
timeDelayFollow=IntVar()
totPost=IntVar()
# data
tk.Spinbox(root,from_= 1, to = 99999.0,width=4, increment=10,textvariable=minFollowData).grid(row=0,column=1)
tk.Spinbox(root,from_= 1, to = 99999.0,width=4, increment=10,textvariable=maxFollowData).grid(row=0,column=3)
tk.Spinbox(root,from_= 1, to = 99999.0,width=4, increment=10,textvariable=minFollowingData).grid(row=1,column=1)
tk.Spinbox(root,from_= 1, to = 99999.0,width=4, increment=10,textvariable=maxFollowingData).grid(row=1,column=3)
tk.Spinbox(root,from_= 1, to = 99999.0,width=4, increment=10,textvariable=timeDelayFollow).grid(row=2,column=1)
tk.Spinbox(root,from_= 1, to = 99999.0,width=4, increment=10,textvariable=totalIds).grid(row=2,column=3)
tk.Spinbox(root,from_= 1, to = 99999.0,width=4, increment=10,textvariable=totPost).grid(row=3,column=1)




Button(text="Run Bot", command=run_All).grid(row=5,column=2)
root.mainloop()
