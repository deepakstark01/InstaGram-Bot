import os
import random
import datetime
from tkinter import *
import tkinter as tk
import shutil
import sys
import time
import pyautogui
import requests
import xlsxwriter
import time
import selenium
from selenium import webdriver
from pynput.keyboard import Controller, Key
from selenium.webdriver.common.keys import Keys
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import pyautogui
from datetime import datetime
from datetime import timedelta

def run_chrome():
    global driver
    chrome_options = webdriver.ChromeOptions()
    # chrome_options.add_argument(r'--user-data-dir=C:\Users\Deepak\AppData\Local\Google\Chrome\User Data')
    chrome_options.add_argument(r'--user-data-dir=C:\Users\mikel\AppData\Local\Google\Chrome\User Data')
    #chrome_options.add_argument(r'--incognito')
    driver = webdriver.Chrome(executable_path="C:\chromedriver.exe", options=chrome_options)
    print("work")
    url="https://www.instagram.com/"
    driver.get(url)
try:
    run_chrome()
except Exception:
    run_chrome()
file = "userData.xlsx"
wb_obj = openpyxl.load_workbook(file)
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
source = "userData.xlsx"
copyNewFileTemp ="TempUser.xlsx"
shutil.copy(source, copyNewFileTemp)
def likeThePost():
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//body//div//div[contains(@class,'_aa')]//div//div//div[1]//div[1]//a[1]"))).click()
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH,"(//*[name()='svg'][@aria-label='Like'])[2]"))).click()
    nextbutton ="//div[contains(@class,'_aaqg _aaqh')]//span"
    maxxlike=maxLike.get()
    minLike=mLike.get()
    for i in range(0,random.randrange(minLike, maxxlike)):
        time.sleep(random.randrange(2, 5))
        try:
            WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH,"(//*[name()='svg'][@aria-label='Like'])[2]"))).click()
            print("Liked ")
        except Exception:
            pass
        try:
            time.sleep(random.randrange(2, 5))
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, nextbutton))).click()
        except Exception:
            break
            pass

def activityFollowUnfollow():
    outWorkbook = xlsxwriter.Workbook("userData.xlsx")
    outSheet = outWorkbook.add_worksheet()
    outSheet.write("A1", "UserIds")
    outSheet.write("B1", "Follower")
    outSheet.write("C1", "Following")
    outSheet.write("D1", "DateFollow")
    outSheet.write("E1", "Follow Done")
    checkFollow = followUSer.get()
    checkUnFollow=unfollowUser.get()
    storycheck = storyViweUser.get()
    numfollowgap = userfol.get()
    numfollow=1
    mint1=mintimepuase.get()
    maxt2=maxtimepasue.get()
    for i in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)
        Insta_Id = cell_obj.value
        followerNumberID = sheet_obj.cell(row=i, column=2).value
        followinNumberID=sheet_obj.cell(row=i, column=3).value
        dateprivous=sheet_obj.cell(row=i, column=4).value
        Insta_Id_Urls = "https://www.instagram.com/" + Insta_Id
        print(str(i) + ":" + Insta_Id_Urls)
        time.sleep(random.randrange(4, 12))
        driver.get(Insta_Id_Urls)
        if (checkFollow):
            print(checkFollow, "true")
            try:
                if(numfollow>numfollowgap):
                    time.sleep(random.randrange(mint1, maxt2))
                    numfollow=1
                time.sleep(random.randrange(5, 15))
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//header//div[contains(text(),'Follow')]"))).click()
                numfollow+=1
                date_current=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                dateText=str(date_current)
                outSheet.write(i-1, 0,Insta_Id)
                outSheet.write(i-1, 1,followerNumberID)
                outSheet.write(i-1, 2,followinNumberID)
                outSheet.write(i-1, 3,dateText)
                outSheet.write(i-1, 4,"Done")
                print(dateText)
                likeThePost()
            except Exception:
                outSheet.write(i - 1, 0, Insta_Id)
                outSheet.write(i - 1, 1, followerNumberID)
                outSheet.write(i - 1, 2, followinNumberID)
                print("already following")
                outSheet.write(i-1, 4,"already following")
        elif(checkUnFollow):
            dateVal=str(dateprivous)
            print(dateVal)
            datetime_object=datetime.strptime(dateVal, '%Y-%m-%d %H:%M:%S')
            today = datetime.now()
            dayval = today - datetime_object  # 09-06-2022 - 08-06-2022 = 1 day
            currntday=int(dayval.days) # 1
            print(currntday) # 1
            daysTofunfollow=daysForUnfollow.get() # user by enterd
            if(daysTofunfollow<=currntday):
                try:
                    time.sleep(random.randrange(5, 15))
                    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH,
                                                                                    "//span//span[1]//button[1]"))).click()
                    try:
                        driver.switch_to_active_element()
                        time.sleep(random.randrange(5, 15))
                        WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.XPATH, "//button[normalize-space()='Unfollow']"))).click()
                        time.sleep(random.randrange(5,15))
                    except Exception:
                        print("not follow")
                except Exception:
                    print("not follow skip")
                print("done unfollow")

        elif(storycheck):
            storyviwe()
    # os.remove("userData.xlsx")
    outWorkbook.close()

def storyviwe():
    print("story")
    driver.get("https://www.instagram.com/")
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "OE3OK"))).click()


def run_All():
    # run_chrome()

    activityFollowUnfollow()


root = Tk()
root.geometry("900x400")
root.title("Author Darakhsha_Deepak")
root['background'] = '#008000'
off_color = "red"
on_color = "green"
# credential
followUSer =IntVar()
unfollowUser =IntVar()
storyViweUser =IntVar()
daysForUnfollow=IntVar()
mLike=IntVar()
maxLike=IntVar()
userfol=IntVar()
mintimepuase=IntVar()
maxtimepasue=IntVar()
Label(root, text="Follow for check", font="Roboto,14,bold", bg="#008000", fg="black").grid(row=0,column=0)
Checkbutton(root, text="Follow", variable=followUSer).grid(row=0,column=1)

Label(root, text="Unfollow for check", font="Roboto,14,bold", bg="#008000", fg="black").grid(row=0,column=2)
Checkbutton(root, text="Unfollow for check", variable=unfollowUser).grid(row=0,column=3)
Label(root, text="Story View", font="Roboto,14,bold", bg="#008000", fg="black").grid(row=2,column=0)
Checkbutton(root, text="Story Viwe for check", variable=storyViweUser).grid(row=2,column=1)
Label(root, text="Days for unfollow", font="Roboto,14,bold", bg="#008000", fg="black").grid(row=3,column=0)
tk.Spinbox(root,from_= 1, to = 99999.0,width=4, increment=10,textvariable=daysForUnfollow).grid(row=3,column=1)
Label(root, text="Min like", font="Roboto,14,bold", bg="#008000", fg="black").grid(row=4,column=0)
tk.Spinbox(root,from_= 1, to = 99999.0,width=4, increment=10,textvariable=mLike).grid(row=4,column=1)
Label(root, text="Max like", font="Roboto,14,bold", bg="#008000", fg="black").grid(row=4,column=2)
tk.Spinbox(root,from_= 1, to = 99999.0,width=4, increment=10,textvariable=maxLike).grid(row=4,column=3)
Label(root, text="USer follow pause gap", font="Roboto,14,bold", bg="#008000", fg="black").grid(row=5,column=2)
tk.Spinbox(root,from_= 1, to = 99999.0,width=4, increment=10,textvariable=userfol).grid(row=5,column=3)
Label(root, text="min time sec", font="Roboto,14,bold", bg="#008000", fg="black").grid(row=6,column=2)
tk.Spinbox(root,from_= 1, to = 99999.0,width=4, increment=10,textvariable=mintimepuase).grid(row=6,column=3)
Label(root, text="max time sec", font="Roboto,14,bold", bg="#008000", fg="black").grid(row=6,column=4)
tk.Spinbox(root,from_= 1, to = 99999.0,width=4, increment=10,textvariable=maxtimepasue).grid(row=6,column=5)

# checkbox for private , photo
# daily follow, unfollow
#like follow


# password.grid(row=1)

# entry


# data





Button(text="Run Bot", command=run_All).grid(row=5,column=0)
root.mainloop()
